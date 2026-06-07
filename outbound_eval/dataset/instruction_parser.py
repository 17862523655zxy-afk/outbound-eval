"""Parser for natural-language task instructions.

Accepts the 6-section format used in the hackathon prompt
(Role / Task / Opening Line / Call Flow / FAQ / Constraints) and produces
a fully populated :class:`EvaluationTask` that can be saved to YAML.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

import yaml
from pydantic import BaseModel, Field

from outbound_eval.dataset.task import (
    DifficultyLevel,
    EvaluationTask,
    FailureCondition,
    SuccessCondition,
)


# ---------------------------------------------------------------------------
# Section model
# ---------------------------------------------------------------------------


class FlowStep(BaseModel):
    """A single step inside the # Call Flow section."""

    step: int
    description: str
    is_p0: bool = False  # flagged when the description contains "必须/务必/一定要"


class FAQItem(BaseModel):
    """A single bullet inside the # Knowledge Points (FAQ) section."""

    index: int
    content: str


class ParsedInstruction(BaseModel):
    """Structured result of parsing a task instruction text.

    `errors` collects non-fatal parsing problems so the caller can decide
    whether to proceed or surface them to the user.
    """

    task_name: str = ""
    role: str = ""
    description: str = ""
    opening_line: str = ""
    variables: dict[str, str] = Field(default_factory=dict)
    flow_steps: list[FlowStep] = Field(default_factory=list)
    faq_items: list[FAQItem] = Field(default_factory=list)
    constraints: list[str] = Field(default_factory=list)
    errors: list[str] = Field(default_factory=list)


# ---------------------------------------------------------------------------
# Section splitting
# ---------------------------------------------------------------------------

_SECTION_RE = re.compile(
    r"^#\s*(Role|Task|Opening\s+Line|Call\s+Flow|Knowledge\s+Points\s*\(FAQ\)|Constraints)\s*$",
    re.IGNORECASE | re.MULTILINE,
)

_SECTION_ALIASES = {
    "role": "role",
    "task": "task",
    "opening line": "opening_line",
    "opening_line": "opening_line",
    "call flow": "flow",
    "call_flow": "flow",
    "knowledge points (faq)": "faq",
    "knowledge points": "faq",
    "faq": "faq",
    "constraints": "constraints",
}

_VAR_RE = re.compile(r"\$\{([^}]+)\}")
_BOLD_VAR_RE = re.compile(r"\*\*([A-Z$])[^*]*\*\*")
_FLOW_RE = re.compile(r"^\s*(\d+)\.\s+(.+)$", re.MULTILINE)
_FAQ_RE = re.compile(r"^\s*-\s+(.+)$", re.MULTILINE)
_P0_HINTS = ("必须", "务必", "一定要", "不得", "不可")


def _split_sections(text: str) -> dict[str, str]:
    """Split raw text into a `{section_key: body}` dict.

    Unknown sections are ignored. The text before the first header (if any)
    is dropped — instructions should always start with `# Role`.
    """
    matches = list(_SECTION_RE.finditer(text))
    if not matches:
        return {}

    sections: dict[str, str] = {}
    for i, m in enumerate(matches):
        header = m.group(1).strip().lower()
        key = _SECTION_ALIASES.get(header)
        if not key:
            continue
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        sections[key] = text[start:end].strip()

    return sections


# ---------------------------------------------------------------------------
# Per-section parsing helpers
# ---------------------------------------------------------------------------


def _parse_role(body: str) -> str:
    return body.strip()


def _parse_task(body: str) -> tuple[str, str]:
    """Return (task_name, description) from the `# Task` section.

    First non-empty line is treated as the task name; the rest is description.
    """
    lines = [l.rstrip() for l in body.splitlines() if l.strip()]
    if not lines:
        return "", ""
    name = lines[0]
    desc_lines = [l.lstrip() for l in lines[1:]] if len(lines) > 1 else []
    return name, "\n".join(desc_lines).strip()


def _parse_opening_line(body: str) -> tuple[str, dict[str, str]]:
    """Extract `${var}` placeholders and `**X**`-style bold variables."""
    text = body.strip()
    variables: dict[str, str] = {}

    for m in _VAR_RE.finditer(text):
        name = m.group(1).strip()
        if name and name not in variables:
            variables[name] = ""

    for m in _BOLD_VAR_RE.finditer(text):
        sym = m.group(1)
        if sym not in variables:
            variables[sym] = ""

    return text, variables


def _parse_flow(body: str) -> list[FlowStep]:
    """Parse `# Call Flow` numbered items into FlowStep objects."""
    steps: list[FlowStep] = []
    for m in _FLOW_RE.finditer(body):
        idx = int(m.group(1))
        text = m.group(2).strip()
        is_p0 = any(hint in text for hint in _P0_HINTS)
        steps.append(FlowStep(step=idx, description=text, is_p0=is_p0))
    return steps


def _parse_faq(body: str) -> list[FAQItem]:
    """Parse `# Knowledge Points (FAQ)` bullet list."""
    items: list[FAQItem] = []
    for i, m in enumerate(_FAQ_RE.finditer(body), start=1):
        items.append(FAQItem(index=i, content=m.group(1).strip()))
    return items


def _parse_constraints(body: str) -> list[str]:
    """Constraints section is a bullet list of free-form rules."""
    return [m.group(1).strip() for m in _FAQ_RE.finditer(body)]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def parse_instruction(text: str) -> ParsedInstruction:
    """Parse a multi-section task instruction into a structured form.

    Never raises; populates ``errors`` for non-fatal issues instead.
    """
    errors: list[str] = []
    sections = _split_sections(text)

    if not sections:
        errors.append("未识别到任何 section 标题（需要以 `# Role` 开头）")
        return ParsedInstruction(errors=errors)

    if "role" not in sections:
        errors.append("缺少 `# Role` section")

    role = _parse_role(sections.get("role", ""))
    task_name, description = _parse_task(sections.get("task", ""))
    opening_line, variables = _parse_opening_line(sections.get("opening_line", ""))
    flow_steps = _parse_flow(sections.get("flow", ""))
    faq_items = _parse_faq(sections.get("faq", ""))
    constraints = _parse_constraints(sections.get("constraints", ""))

    if not flow_steps:
        errors.append("# Call Flow section 未解析出任何步骤（请使用 `数字. ` 开头）")
    if not faq_items:
        errors.append("# Knowledge Points (FAQ) section 未解析出任何条目（请使用 `- ` 开头）")

    return ParsedInstruction(
        task_name=task_name,
        role=role,
        description=description,
        opening_line=opening_line,
        variables=variables,
        flow_steps=flow_steps,
        faq_items=faq_items,
        constraints=constraints,
        errors=errors,
    )


def _infer_skill_name(role: str, task_name: str) -> str:
    """Best-effort skill name inference from the role + task.

    Heuristics: take the first distinct Chinese proper noun, or fall back
    to the task_id prefix. Always lowercased + ASCII for safety.
    """
    text = f"{role} {task_name}"
    # Look for known brand-like tokens
    for keyword in ("飞毛腿", "美团", "外卖", "直播课", "课程", "站长"):
        if keyword in text:
            mapping = {
                "飞毛腿": "feimaotui",
                "美团": "meituan",
                "外卖": "waimai",
                "直播课": "live_course",
                "课程": "course",
                "站长": "station_master",
            }
            return mapping[keyword]
    # Fall back to ASCII slug of the first Hanzi run
    m = re.search(r"[\u4e00-\u9fff]+", task_name or role)
    if m:
        return m.group(0)
    return "default"


def to_evaluation_task(parsed: ParsedInstruction, task_id: str) -> EvaluationTask:
    """Convert a ParsedInstruction into a fully populated EvaluationTask."""
    name = parsed.task_name or task_id
    description = parsed.description or name

    skill_name = _infer_skill_name(parsed.role, name)

    # Build success_criteria: one entry per flow step (P0 for must-words),
    # plus one FAQ-coverage entry per FAQ item.
    success: list[SuccessCondition] = []
    for s in parsed.flow_steps:
        # short, label-like name: first 12 chars of description
        label = s.description[:12].rstrip("，,。. ") or f"流程步骤 {s.step}"
        success.append(SuccessCondition(
            condition_id=f"flow_step_{s.step}",
            name=label,
            description=s.description,
            priority="P0" if s.is_p0 else "P1",
            check_type="llm",
            check_config={"required_keywords": [], "flow_step": s.step},
            weight=1.0,
        ))

    for faq in parsed.faq_items:
        # Short, FAQ-like label: first 16 chars
        label = ("FAQ: " + faq.content[:14]).rstrip("，,。. ")
        success.append(SuccessCondition(
            condition_id=f"faq_coverage_{faq.index}",
            name=label,
            description=(
                f"用户提问时，agent 应正确回答以下 FAQ 内容：{faq.content}"
            ),
            priority="P1",
            check_type="llm",
            check_config={"required_keywords": [], "faq_index": faq.index},
            weight=0.8,
        ))

    # Minimal failure criteria: do not fabricate, do not skip if no constraints
    failure: list[FailureCondition] = []
    for i, c in enumerate(parsed.constraints, start=1):
        failure.append(FailureCondition(
            condition_id=f"constraint_{i}",
            name=f"违反约束 {i}",
            description=c,
            priority="P1",
            check_type="llm",
            check_config={"constraint": c},
            penalty_weight=1.0,
        ))

    return EvaluationTask(
        task_id=task_id,
        name=name,
        description=description,
        skill_name=skill_name,
        variables=parsed.variables or {},
        difficulty=DifficultyLevel.MEDIUM,
        success_criteria=success,
        failure_criteria=failure,
        pass_threshold=0.7,
        expected_outcome={"task_completed": True, "user_commitment": "confirm"},
        injected_events=[],
    )


def save_task_yaml(task: EvaluationTask, output_path: Path) -> Path:
    """Serialize an EvaluationTask to YAML at ``output_path``.

    Creates parent directories as needed. Returns the resolved path.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        yaml.safe_dump(
            task.model_dump(mode="json"),
            allow_unicode=True,
            default_flow_style=False,
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    return output_path


def parse_and_save(
    text: str,
    task_id: str,
    output_path: Optional[Path] = None,
) -> tuple[ParsedInstruction, EvaluationTask, Optional[Path]]:
    """Convenience: parse → convert → optionally save to disk.

    Returns (parsed, task, saved_path). `saved_path` is None when no
    `output_path` was provided.
    """
    parsed = parse_instruction(text)
    task = to_evaluation_task(parsed, task_id=task_id)
    saved = save_task_yaml(task, output_path) if output_path else None
    return parsed, task, saved


# ---------------------------------------------------------------------------
# Excel loading
# ---------------------------------------------------------------------------


def _extract_excel_text(path: Path, row_index: int = 0) -> str:
    """Read Sheet1 of an .xlsx file and return its content as plain text.

    The first column is treated as a label/ID and skipped; subsequent
    columns are joined with newlines. Only one row is returned
    (``row_index`` selects it, defaulting to the first content row).

    The hackathon `命题二` sample has one task per row — picking a single
    row avoids cross-task bleed when the sheet contains multiple examples.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "需要安装 openpyxl 才能读 .xlsx 文件：pip install openpyxl"
        ) from e

    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    try:
        sheet = wb.active if "Sheet1" not in wb.sheetnames else wb["Sheet1"]
        lines: list[str] = []
        header_skipped = False
        content_row_index = 0
        for row in sheet.iter_rows(values_only=True):
            cells = [("" if c is None else str(c)).strip() for c in row]
            if not any(cells):
                continue
            if not header_skipped and any(
                h in cells[0] for h in ("id", "ID", "编号", "序号")
            ):
                header_skipped = True
                continue
            header_skipped = True
            if content_row_index < row_index:
                content_row_index += 1
                continue
            instruction = "\n".join(c for c in cells[1:] if c)
            if instruction:
                lines.append(instruction)
            break  # only the requested row
        return "\n".join(lines)
    finally:
        wb.close()


def parse_excel_file(
    path: Path,
    task_id: str,
    output_path: Optional[Path] = None,
    row_index: int = 0,
) -> tuple[ParsedInstruction, EvaluationTask, Optional[Path], str]:
    """Parse an .xlsx task instruction file and (optionally) save YAML.

    Returns ``(parsed, task, saved_path, raw_text)``. ``raw_text`` is the
    extracted markdown-style instruction text, useful for debugging.
    Use ``row_index`` to pick which content row to read.
    """
    path = Path(path)
    text = _extract_excel_text(path, row_index=row_index)
    parsed, task, saved = parse_and_save(text, task_id, output_path)
    return parsed, task, saved, text
